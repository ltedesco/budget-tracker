#!/usr/bin/env python3
"""Build a budget-tracker seed document from an annual-budget workbook.

    python3 scripts/seed-from-xlsx.py Annual_budget_2026.xlsx 2026 > seed.json

The output is a full document ready for Import in the app. It contains real
amounts, so it is gitignored here and belongs only in the private data repo.

Workbook layout this expects (the standard annual-budget template):
  col A  category name, on the category header row only
  col C  line-item name  ('Monthly totals:' marks the header row)
  cols D..O  the twelve months
A blank cell stays blank (null) rather than becoming 0 — "not recorded" and
"zero" are different things, and the app keeps them apart.
"""
import json
import sys
import uuid
import openpyxl

NAME_COL, FIRST_MONTH_COL = 2, 3  # zero-based


def new_id():
    return uuid.uuid4().hex[:8]


def cell(v):
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def read_sheet(ws, kind, at, categories, items):
    current = None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        vals = [c.value for c in row]
        head = str(vals[0]).strip() if vals[0] is not None else ""
        name = str(vals[NAME_COL]).strip() if len(vals) > NAME_COL and vals[NAME_COL] is not None else ""

        if head:  # a category header row
            current = {
                "id": new_id(),
                "kind": kind,
                "name": head,
                "order": len(categories),
                "updatedAt": at,
            }
            categories.append(current)
            continue

        if not current or not name or name.lower().startswith("monthly totals"):
            continue

        months = [cell(vals[FIRST_MONTH_COL + i]) if FIRST_MONTH_COL + i < len(vals) else None
                  for i in range(12)]

        items.append({
            "id": new_id(),
            "categoryId": current["id"],
            "name": name,
            "order": sum(1 for it in items if it["categoryId"] == current["id"]),
            "planned": months,
            "actual": [None] * 12,
            "fieldsAt": {},
            "baseAt": at,
            "updatedAt": at,
        })


def main():
    path = sys.argv[1]
    year = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    at = f"{year}-01-01T00:00:00.000Z"

    wb = openpyxl.load_workbook(path, data_only=True)
    categories, items = [], []

    if "Income" in wb.sheetnames:
        read_sheet(wb["Income"], "income", at, categories, items)
    if "Expenses" in wb.sheetnames:
        read_sheet(wb["Expenses"], "expense", at, categories, items)

    starting = 0
    if "Setup" in wb.sheetnames:
        for row in wb["Setup"].iter_rows():
            for i, c in enumerate(row):
                if isinstance(c.value, str) and "starting balance" in c.value.lower():
                    for nxt in row[i + 1:]:
                        if isinstance(nxt.value, (int, float)):
                            starting = float(nxt.value)
                            break

    json.dump({
        "version": 1,
        "year": year,
        "startingBalance": starting,
        "startingBalanceAt": at,
        "categories": categories,
        "items": items,
        "deleted": [],
    }, sys.stdout, indent=2)
    print()


if __name__ == "__main__":
    main()
